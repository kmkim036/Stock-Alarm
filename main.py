import requests
import json
import openpyxl
import datetime


class Enterprise:
    def __init__(self, name, itemcode):
        self.name = name
        self.itemcode = itemcode

    def referAPI(self):
        host = 'http://api.finance.naver.com'
        path = '/service/itemSummary.nhn'
        param = {'itemcode': self.itemcode}
        url = host + path

        res = requests.get(url, params=param)
        if res.status_code != 200:
            print("ERROR:" + "Get stock-info API Failed in " + self.name)
            return

        data = res.json()

        # revise condition for alarm after
        if data["risefall"] > 3:
            sendKakaotalk(self.name)

        self.save(data)

    def save(self, data):
        try:
            wb = openpyxl.load_workbook(
                '/home/kmkim/Projects/git/kmkim036/Stock-Manage/data.xlsx')
            ws = wb[self.name]
        except FileNotFoundError:
            print("ERROR: " + "File 'data.xlsx' open error in " + self.name)
            return

        A_lastrow = 'A' + str(ws.max_row)
        dt_now = datetime.datetime.now()
        if dt_now.date() < ws[A_lastrow].value.date():
            print("ERROR: " + "Date overflow in " + self.name)
            wb.close()
            return

        if dt_now.date() == ws[A_lastrow].value.date():
            ws.delete_rows(ws.max_row)

        risefall = data["risefall"]
        if risefall == 1:
            rf = "상한"
        elif risefall == 2:
            rf = "상승"
        elif risefall == 3:
            rf = "보합"
        elif risefall == 4:
            rf = "하한"
        else:
            rf = "하락"

        ws.append([dt_now.date(), data["marketSum"], data["per"], data["eps"], data["pbr"], data["now"],
                   data["diff"], data["rate"], data["quant"], data["amount"], data["high"], data["low"], rf])
        wb.save('/home/kmkim/Projects/git/kmkim036/Stock-Manage/data.xlsx')
        wb.close()


def sendKakaotalk(ent_name):
    try:
        with open('/home/kmkim/Projects/security.json') as json_file:
            json_data = json.load(json_file)
            KAKAO_TOKEN = json_data["kakao"]["self-send-token"]
    except FileNotFoundError:
        print("ERROR: " + "File 'security.json' open error in " + ent_name)
        return

    header = {"Authorization": 'Bearer ' + KAKAO_TOKEN}
    url = "https://kapi.kakao.com/v2/api/talk/memo/default/send"

    body = {
        "object_type": "text",
        "text": ent_name + " 돔황챠~~",
        "link": {
            "web_url": "https://developers.kakao.com"
        }
    }
    body_json = {"template_object": json.dumps(body)}

    res = requests.post(url, headers=header, data=body_json)
    if res.status_code != 200:
        print("ERROR: " + "post kakao-self-send API in " + ent_name)

    return


e1 = Enterprise("samsung", "005930")
e1.referAPI()

e2 = Enterprise("kakao", "035720")
e2.referAPI()

e3 = Enterprise("SKhynix", "000660")
e3.referAPI()

e4 = Enterprise("naver", "035420")
e4.referAPI()
