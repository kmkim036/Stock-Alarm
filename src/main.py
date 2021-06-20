import sys
import datetime
import json
import openpyxl

import log
import naver_api
import kakao_api
import mail


data_FILE_PATH = '/home/kmkim/Projects/git/kmkim036/Stock-Alarm/data/data.xlsx'


class Enterprise:
    def __init__(self, name, itemcode, purchase_price):
        self.name = name
        self.itemcode = itemcode
        self.purchase_price = purchase_price

    def get_data(self):
        ret = naver_api.get_stock_data(self.name, self.itemcode)
        if ret == -1 or ret == -2:
            return
        else:
            self.save_data(ret)

    def save_data(self, data):
        try:
            wb = openpyxl.load_workbook(data_FILE_PATH)
            ws = wb[self.name]
        except FileNotFoundError:
            msg = "File open error"
            log.record_error(msg, sys._getframe().f_code.co_name)
            return

        lastrow = ws.max_row
        dt_now = datetime.datetime.now()

        if dt_now.date() < ws['A' + str(lastrow)].value.date():
            msg = "Date overflow"
            log.record_error(msg, sys._getframe().f_code.co_name)
            wb.close()
            return

        if dt_now.date() == ws['A' + str(lastrow)].value.date():
            ws.delete_rows(ws.max_row)

        risefall = data["risefall"]
        if risefall == 1:
            rf = "상한"
        elif risefall == 2:
            rf = "상승"
        elif risefall == 3:
            rf = "보합"
        elif risefall == 4:
            rf = "하한"
        else:
            rf = "하락"

        lower_threshold = self.purchase_price * 0.95
        upper_threshold = self.purchase_price * 1.05

        if lastrow < 6:
            check_range = range(2, lastrow + 1)
        else:
            check_range = range(lastrow - 3, lastrow + 1)

        check = 'n'
        for i in check_range:
            value = ws['F' + str(i)].value
            if value < lower_threshold:
                check = 'y_l'
            elif value > upper_threshold:
                check = 'y_u'

        if check != 'y_l' and data['now'] < lower_threshold:
            kakao_api.send_to_me(self.name, 1)
            mail.send_mail(self.name, 1)
            
        if check != 'y_u' and data['now'] > upper_threshold:
            kakao_api.send_to_me(self.name, 2)
            mail.send_mail(self.name, 2)

        ws.append([dt_now.date(), data["marketSum"], data["per"], data["eps"], data["pbr"], data["now"],
                   data["diff"], data["rate"], data["quant"], data["amount"], data["high"], data["low"], rf])
        wb.save(data_FILE_PATH)
        wb.close()


if __name__ == "__main__":
    if kakao_api.refresh_token() < 0:
        sys.exit(1)

    e1 = Enterprise("samsung", "005930", 80600)
    e1.get_data()

    e2 = Enterprise("kakao", "035720", 127000)
    e2.get_data()

    e3 = Enterprise("SKhynix", "000660", 128500)
    e3.get_data()

    e4 = Enterprise("naver", "035420", 367000)
    e4.get_data()
